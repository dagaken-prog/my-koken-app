import openpyxl
import os
import glob

# ファイル名パターン
file_pattern = "01.家裁定期報告*.xlsx"
files = glob.glob(file_pattern)

if not files:
    print("Excel file not found.")
    exit()

target_file = files[0]
print(f"Analyzing: {target_file}")

try:
    wb = openpyxl.load_workbook(target_file, data_only=True)
except Exception as e:
    print(f"Error loading workbook: {e}")
    exit()

output_file = "excel_analysis_report.txt"

with open(output_file, "w", encoding="utf-8") as f:
    f.write(f"# Excel Analysis Report: {os.path.basename(target_file)}\n\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f"## Sheet: {sheet_name}\n")
        f.write(f"Size: {ws.max_row} rows x {ws.max_column} columns\n")
        
        # 結合セル情報
        f.write("### Merged Cells:\n")
        for merged_range in ws.merged_cells.ranges:
            f.write(f"- {merged_range}\n")
        
        # 値のあるセルをピックアップ (先頭100行くらいまで)
        f.write("\n### Cell Values (Non-empty, first 100 rows):\n")
        for row in ws.iter_rows(max_row=100):
            for cell in row:
                if cell.value is not None:
                    val = str(cell.value).strip()
                    if val:
                        # 改行を含む場合は1行にする
                        val_clean = val.replace('\n', '\\n')
                        f.write(f"- {cell.coordinate}: {val_clean}\n")
        f.write("\n" + "="*50 + "\n\n")

print(f"Analysis complete. Report saved to {output_file}")
