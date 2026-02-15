
import io
import openpyxl
from openpyxl.cell.cell import MergedCell
import datetime
from .constants import (
    EXCEL_TEMPLATE_PATH, SHEET_REPORT, SHEET_ASSETS,
    CELL_REPORT_NAME_KANA, CELL_REPORT_NAME, CELL_REPORT_ADDRESS, 
    CELL_REPORT_POSTAL, CELL_REPORT_RESIDENCE, CELL_REPORT_RES_POSTAL,
    CELL_REPORT_PERIOD_START, CELL_REPORT_PERIOD_END, CELL_REPORT_CREATE_DATE,
    CELL_GUARDIAN_ADDRESS, CELL_GUARDIAN_NAME, CELL_GUARDIAN_PHONE,
    CELL_ASSET_NAME, ROW_ASSET_BANK_START,
    CELL_ASSET_CASH_VAL, CELL_ASSET_FACILITY_VAL,
    COL_BANK_NAME, COL_BANK_BRANCH, COL_BANK_TYPE_FUTSU, COL_BANK_TYPE_TEIKI,
    COL_BANK_NUMBER, COL_BANK_DATE, COL_BANK_VALUE, COL_BANK_ADMIN
)

def safe_set_value(ws, coord, value):
    """
    セルに値を設定する（結合セルの場合は左上のみ設定、エラー回避）
    """
    try:
        cell = ws[coord]
        if isinstance(cell, MergedCell):
            # 結合セルの場合、そのセルが左上かどうか判断するのは難しいが、
            # openpyxlのws[coord]は結合セル内の特定セルを指す。
            # 値セットでエラーが出るならスキップ
            pass
        else:
            cell.value = value
    except AttributeError:
        # MergedCellにvalueセットしようとした場合などに発生
        pass
    except Exception:
        # その他のエラー(ReadOnly等)も無視して続行
        pass

def safe_cell_write(ws, row, col, value):
    """
    行・列指定で値を設定
    """
    try:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            pass
        else:
            cell.value = value
    except AttributeError:
        pass
    except Exception:
        pass

def create_periodic_report(person_data, guardian_data, asset_list, activity_list, template_path=EXCEL_TEMPLATE_PATH):
    """
    定期報告書Excelを作成する関数
    """
    try:
        wb = openpyxl.load_workbook(template_path)
    except FileNotFoundError:
        return None, "テンプレートファイルが見つかりません"

    # --- 共通変数 ---
    today = datetime.date.today()
    
    # シート名取得 (スペース等の揺らぎ対策)
    sheet_names = wb.sheetnames
    target_report_sheet = None
    target_assets_sheet = None
    
    for sn in sheet_names:
        if SHEET_REPORT in sn: # 部分一致で探す
            target_report_sheet = sn
        if SHEET_ASSETS in sn:
            target_assets_sheet = sn

    # --- 後見事務報告書シート ---
    if target_report_sheet:
        ws = wb[target_report_sheet]
        
        # 基本情報
        safe_set_value(ws, CELL_REPORT_NAME, person_data.get('氏名', ''))
        # safe_set_value(ws, CELL_REPORT_NAME_KANA, person_data.get('ｼﾒｲ', '')) 
        
        safe_set_value(ws, CELL_REPORT_ADDRESS, person_data.get('住所', ''))
        safe_set_value(ws, CELL_REPORT_POSTAL, person_data.get('〒', ''))
        
        safe_set_value(ws, CELL_REPORT_RESIDENCE, person_data.get('居所', ''))
        # safe_set_value(ws, CELL_REPORT_RES_POSTAL, "") 

        # 後見人情報
        safe_set_value(ws, CELL_GUARDIAN_ADDRESS, guardian_data.get('住所', ''))
        safe_set_value(ws, CELL_GUARDIAN_NAME, guardian_data.get('氏名', ''))
        safe_set_value(ws, CELL_GUARDIAN_PHONE, guardian_data.get('連絡先電話番号', ''))

        # 報告期間と作成日
        today = datetime.date.today()
        report_month = person_data.get('家裁報告月')
        if report_month:
            try:
                rm = int(str(report_month).replace('月', ''))
                target_date = datetime.date(today.year, rm, 1)
                if target_date < today:
                     start_date = datetime.date(today.year - 1, rm, 1)
                     end_date = target_date - datetime.timedelta(days=1)
                else:
                     start_date = datetime.date(today.year - 1, rm, 1)
                     end_date = datetime.date(today.year, rm, 1) - datetime.timedelta(days=1)

                safe_set_value(ws, CELL_REPORT_PERIOD_START, start_date)
                safe_set_value(ws, CELL_REPORT_PERIOD_END, end_date)
            except:
                pass
        
        safe_set_value(ws, CELL_REPORT_CREATE_DATE, today)

    # --- 財産目録シート ---
    if target_assets_sheet:
        ws_ast = wb[target_assets_sheet]
        safe_set_value(ws_ast, CELL_ASSET_NAME, person_data.get('氏名', ''))
        
        # 既存データのクリア (行 25〜39 の範囲を想定)
        # B~AF列までクリア。結合セル対策
        for r in range(ROW_ASSET_BANK_START, 40): # 39行目まで念のため
            for c in range(2, 33): # B(2) ~ AF(32)
                safe_cell_write(ws_ast, r, c, None) # Noneを設定してクリア

        # --- 預貯金 ---
        real_banks = [a for a in asset_list if a.get('財産種別') == '預貯金']
        
        row_idx = ROW_ASSET_BANK_START
        for i, bank in enumerate(real_banks):
            if row_idx > 36: break # 現金欄(37)に被らないように
            
            # 金融機関名
            safe_cell_write(ws_ast, row_idx, COL_BANK_NAME, bank.get('名称・機関名', ''))
            # 支店名
            safe_cell_write(ws_ast, row_idx, COL_BANK_BRANCH, bank.get('支店・詳細', ''))
            
            # 種別チェック
            detail = str(bank.get('支店・詳細', '')) + str(bank.get('備考', ''))
            if '定期' in detail or '定額' in detail:
                 safe_cell_write(ws_ast, row_idx, COL_BANK_TYPE_TEIKI, "■")
                 safe_cell_write(ws_ast, row_idx, COL_BANK_TYPE_FUTSU, "□")
            else:
                 safe_cell_write(ws_ast, row_idx, COL_BANK_TYPE_FUTSU, "■")
                 safe_cell_write(ws_ast, row_idx, COL_BANK_TYPE_TEIKI, "□")

            # 口座番号
            safe_cell_write(ws_ast, row_idx, COL_BANK_NUMBER, bank.get('口座番号・記号', ''))
            
            # 最終確認日
            u_date = bank.get('更新日')
            safe_cell_write(ws_ast, row_idx, COL_BANK_DATE, u_date if u_date else today)

            # 残高
            try:
                val = int(float(bank.get('評価額・残高', 0)))
                safe_cell_write(ws_ast, row_idx, COL_BANK_VALUE, val)
            except:
                safe_cell_write(ws_ast, row_idx, COL_BANK_VALUE, bank.get('評価額・残高', 0))
            
            # 管理者
            safe_cell_write(ws_ast, row_idx, COL_BANK_ADMIN, "成年後見人")

            row_idx += 2 

        # --- 現金 ---
        cash_items = [a for a in asset_list if a.get('財産種別') == '現金']
        cash_total = 0
        for c in cash_items:
            try:
                cash_total += int(float(c.get('評価額・残高', 0)))
            except:
                pass
        
        # システムに現金データがあれば書き込む (無ければ、クリア済みなので空欄になる)
        # ただし、テンプレートの元の値は「148871」だった。クリア済みなので、0なら0と書くか、空にするか。
        # ユーザー要望：「システムのデータが空欄であれば、エクセル出力も空欄となるように」
        # asset_list自体に現金が無ければ空欄。あれば合計値。
        if cash_items:
             safe_set_value(ws_ast, CELL_ASSET_CASH_VAL, cash_total)
        
        # --- 施設等預入金 ---
        # データ種別として「施設等預入金」があるか不明だが、備考に「施設」を含むもの等のロジックも考えられるが、
        # 現状は種別マスタ依存。もし種別が「その他」で名称が「施設預り金」などの場合など。
        # ここでは単純に種別が「施設等預入金」または名称に「施設」かつ種別「その他」を拾うなど
        fac_items = [a for a in asset_list if a.get('財産種別') == '施設等預入金' or (a.get('財産種別') == 'その他' and '施設' in str(a.get('名称・機関名', '')))]
        if fac_items:
            fac_total = 0
            for f in fac_items:
                try: fac_total += int(float(f.get('評価額・残高', 0)))
                except: pass
            safe_set_value(ws_ast, CELL_ASSET_FACILITY_VAL, fac_total)

    # 保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, None
